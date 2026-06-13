import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Requirements"

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
todo_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

headers = ["#", "Module", "Feature", "Detail", "Status"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal="center"); c.border = thin_border

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 28
ws.column_dimensions["D"].width = 55
ws.column_dimensions["E"].width = 10

items = [
("Project","Create/Switch/Close","Isolated folders per project","Done"),
("Project","Project tree","Tree view, double-click to open","Done"),
("Project","Auto-detect classes","Extract classes from JSON imports","Done"),
("Project","Ignore region","__ignore__ class for masked-out areas","Done"),
("Data","Image import","BMP/PNG/JPG with CJK paths","Done"),
("Data","JSON import/export","LabelMe format, auto-save, cross-tool","Done"),
("Data","Batch JSON import","Import all JSONs from a folder","Done"),
("Data","Train/Val/Test split","3-state split, right-click batch set","Done"),
("Data","Auto split","Auto 70/20/10, K-fold CV","Done"),
("Data","Image filter","Search by name/class, filter by split status","Done"),
("Data","Annotated filter","Combined filter: annotated + train set","Done"),
("Data","Prediction filter","Filter by pred block count >=/=/<","Done"),
("Data","Recent sort","Sort by JSON modification date","Done"),
("Data","Image preview","Right-click preview before assigning split","Done"),
("Annotation","Pan mode","Default pan to avoid accidental marks","Done"),
("Annotation","Brush","Adjustable brush size, multi-class","Done"),
("Annotation","Eraser","Erase annotated regions","Done"),
("Annotation","Polygon","Draw closed polygon regions","Done"),
("Annotation","Line","Line annotation tool","Done"),
("Annotation","Undo","Undo annotation actions","Done"),
("Annotation","Class switch","Left panel class list, click to switch","Done"),
("Annotation","Auto-save","Auto-save JSON after each edit","Done"),
("Annotation","Zoom/Pan","Scroll zoom, middle-button pan","Done"),
("Annotation","Fit zoom","Auto fit image to 99% canvas","Done"),
("Annotation","Class labels","Labels at top-left of each annotation block","Done"),
("Annotation","Hatch overlay","Diagonal hatch pattern vs solid prediction","Done"),
("Annotation","Ann stats table","Per-block: class/area/perimeter/radius/gray","Done"),
("Annotation","Min area filter","Delete blocks below pixel area threshold","Done"),
("Inference","Single inference","Run model on current image","Done"),
("Inference","Batch inference","Infer all images, save mask + overlay","Done"),
("Inference","Progress display","Real-time progress + per-image time","Done"),
("Inference","Stop inference","Stop mid-batch","Done"),
("Inference","Model select","Dropdown: best_model/last_model","Done"),
("Inference","Backend select","PyTorch/TorchScript/ONNX","Done"),
("Inference","Tiled toggle","Tiled vs full-image inference","Done"),
("Inference","Scale slider","25%-100% downsampling for speed","Done"),
("Inference","C++ backend","C++ ONNX Runtime inference","Todo"),
("Display","Annotation overlay Ctrl+Space","Hatch pattern annotation overlay","Done"),
("Display","Prediction overlay Space","Semi-transparent solid fill + labels","Done"),
("Display","Hover tooltip","Hover shows class/area/gray/radius/IoU","Done"),
("Display","Result stats table","Per-block: class/score/area/perimeter/radius/gray/IoU","Done"),
("Display","Maximize canvas","Backtick key collapses all panels","Done"),
("Display","Panel toggle","Ctrl+[ left, Ctrl+] right","Done"),
("Display","Log toggle","Ctrl+L or View menu","Done"),
("Display","Status bar","Resolution/mode/overlay/zoom/brush info","Done"),
("Training","Model selection","DeepLabV3 R50/R101/Mobile, FCN R50/R101, LRASPP","Done"),
("Training","Tiled training","512px sliding window, full-image inference","Done"),
("Training","AMP","Automatic mixed precision","Done"),
("Training","Loss selection","CE/Lovasz/Dice/Focal/CE+Lovasz/CE+Dice","Done"),
("Training","Augmentation","3 levels: flip/color/blur/noise","Done"),
("Training","Class weights","Inverse frequency, bg=1.0","Done"),
("Training","K-fold CV","1-10 fold for stability assessment","Done"),
("Training","Resume","Resume from last checkpoint","Done"),
("Training","Multi-worker","num_workers=4 + pin_memory, 3-4x speed","Done"),
("Training","Batch progress","Real-time batch loss","Done"),
("Training","Per-class IoU","Worst 3 classes each epoch","Done"),
("Training","Loss curve","Matplotlib real-time plot","Done"),
("Training","Stop training","Mid-training stop, GPU release","Done"),
("Training","Best model","Auto-save best mIoU model","Done"),
("Training","Settings persist","Training params saved across sessions","Done"),
("Training","UI settings persist","Tiled/scale UI state saved","Done"),
("Training","Label smoothing","Label smoothing for generalization","Todo"),
("Training","LR scheduler","Cosine/Warmup/ReduceLROnPlateau","Todo"),
("Training","Multi-GPU","DataParallel/DDP distributed training","Todo"),
("Training","ETA","Per-epoch time + remaining estimate","Todo"),
("Eval","mIoU","Validate mIoU + per-class IoU","Done"),
("Eval","Model mgmt","Multi-model list for comparison","Done"),
("Eval","Loss tracking","Train/val loss history","Done"),
("Eval","Confusion matrix","Pixel-level confusion matrix","Todo"),
("Eval","PR curve","Per-class Precision-Recall curve","Todo"),
("Deploy","Export ONNX","ONNX export for C++ inference","Done"),
("Deploy","Export TorchScript","TorchScript export","Done"),
("Deploy","TensorRT","TensorRT inference engine","Todo"),
("Deploy","Model encryption","Encrypt model files","Todo"),
("Collab","Multi-user","Multi-user annotation sync","Todo"),
("Collab","Review","Annotation quality review","Todo"),
("Collab","Versioning","Annotation history/rollback","Todo"),
("System","Restart","One-click app restart","Done"),
("System","Log panel","Resizable log for training/inference","Done"),
("System","GPU monitor","Real-time GPU memory display","Todo"),
("System","Dark theme","Global dark theme","Todo"),
]

row = 2
for i, (mod, feat, detail, status) in enumerate(items, 1):
    ws.cell(row=row, column=1, value=i).border = thin_border
    ws.cell(row=row, column=2, value=mod).border = thin_border
    ws.cell(row=row, column=3, value=feat).border = thin_border
    ws.cell(row=row, column=4, value=detail).border = thin_border
    c = ws.cell(row=row, column=5, value=status)
    c.border = thin_border; c.alignment = Alignment(horizontal="center")
    c.fill = done_fill if status == "Done" else todo_fill
    row += 1

ws.auto_filter.ref = f"A1:E{row-1}"
ws.freeze_panes = "A2"

out = r"C:\Users\Administrator\Desktop\AI_Training_Platform_Requirements.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total: {len(items)} | Done: {sum(1 for _,_,_,_,s in items if s=='Done')} | Todo: {sum(1 for _,_,_,_,s in items if s=='Todo')}")
