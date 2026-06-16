"""Generate PRD xlsx for 图像分类 module, matching original Requirements table format."""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = os.path.join(os.path.dirname(__file__), "PRD_图像分类模块.xlsx")

# --- Data rows matching original template: 序号, 模块, 功能, 详细说明, 状态 ---
rows = [
    # 工程管理
    ("C01", "工程管理", "工程任务类型扩展", 'project.json 中 task_type 新增 "图像分类" 选项', "待实现"),
    ("C02", "工程管理", "分类工程目录结构", "images/ + annotations/(存class_labels.json) + models/ + outputs/", "待实现"),
    ("C03", "工程管理", "任务类型切换", "工程级别切换三种任务类型（分割/检测/分类），切换后面板适配", "待实现"),
    # 数据管理
    ("C04", "数据管理", "图片导入", "复用现有导入逻辑，支持 BMP/PNG/JPG", "待实现"),
    ("C05", "数据管理", "类别标签管理", 'class_labels.json，支持手动编辑，格式如 {"0":"cat","1":"dog",...}', "待实现"),
    ("C06", "数据管理", "文件夹自动标注", "按子文件夹名自动分配类别（ImageFolder 模式：images/cat/001.jpg→类别cat）", "待实现"),
    ("C07", "数据管理", "训练/验证/测试集划分", "复用现有三态划分 train/val/test", "待实现"),
    ("C08", "数据管理", "自动划分数据集", "复用 Auto Split 70/20/10 + K折交叉验证", "待实现"),
    ("C09", "数据管理", "图片列表筛选", "按名称/类别/数据集筛选", "待实现"),
    ("C10", "数据管理", "分类标注统计", "每类样本数柱状图 + 类别分布饼图 + 标注进度", "待实现"),
    # 训练
    ("C11", "训练", "模型选择", "ResNet18/34/50/101, EfficientNet-B0~B4, MobileNetV3, ViT-B/16", "待实现"),
    ("C12", "训练", "预训练权重", "ImageNet 预训练，自动下载/离线缓存", "待实现"),
    ("C13", "训练", "分类头替换", "自动替换 FC 层适配类别数", "待实现"),
    ("C14", "训练", "训练参数", "epochs / batch_size / lr / optimizer (Adam/SGD/AdamW) / weight_decay", "待实现"),
    ("C15", "训练", "损失函数", "CrossEntropy / LabelSmoothing / FocalLoss", "待实现"),
    ("C16", "训练", "学习率调度", "CosineAnnealing / StepLR / ReduceLROnPlateau", "待实现"),
    ("C17", "训练", "数据增强", "RandomResizedCrop / HorizontalFlip / ColorJitter / RandAugment，3档强度", "待实现"),
    ("C18", "训练", "混合精度AMP", "复用现有 AMP 逻辑，auto_cast 训练加速", "待实现"),
    ("C19", "训练", "断点续训", "Resume 从 last_model.pth 恢复，保留优化器状态", "待实现"),
    ("C20", "训练", "多进程数据加载", "复用 num_workers 逻辑，3-4x 训练提速", "待实现"),
    ("C21", "训练", "批次进度+指标", "实时 batch loss + Top-1/Top-5 Accuracy + 每类 Precision/Recall/F1", "待实现"),
    ("C22", "训练", "最佳模型保存", "自动保存最高 Val Accuracy 模型", "待实现"),
    ("C23", "训练", "K折交叉验证", "复用现有 K-fold 逻辑，1-10 折可选", "待实现"),
    ("C24", "训练", "训练停止+显存释放", "中途停止训练，释放 GPU 显存", "待实现"),
    ("C25", "训练", "类别权重", "自动逆频率权重处理不均衡数据", "待实现"),
    ("C26", "训练", "混淆矩阵", "训练完成后自动生成 + matplotlib/seaborn 热力图", "待实现"),
    # 推理
    ("C27", "推理", "单图推理", "对当前图片分类，显示 Top-5 类别+置信度", "待实现"),
    ("C28", "推理", "批量推理", "遍历全部图片推理，输出 CSV/JSON 结果到 outputs", "待实现"),
    ("C29", "推理", "推理进度+耗时", "复用现有实时进度 + 每张耗时显示", "待实现"),
    ("C30", "推理", "推理停止", "复用现有停止按钮，中途停止批量推理", "待实现"),
    ("C31", "推理", "模型选择", "复用模型下拉框，自动识别分类/检测/分割模型", "待实现"),
    ("C32", "推理", "推理后端", "PyTorch / TorchScript / ONNX Runtime", "待实现"),
    ("C33", "推理", "推理结果可视化", '图片列表新增"预测类别"列，右侧 Top-5 置信度柱状图', "待实现"),
    # 评估
    ("C34", "评估", "Top-1 / Top-5 Accuracy", "验证集整体指标", "待实现"),
    ("C35", "评估", "每类 Precision / Recall / F1", "输出各类别定位短板", "待实现"),
    ("C36", "评估", "混淆矩阵可视化", "seaborn/matplotlib 热力图可视化", "待实现"),
    ("C37", "评估", "模型管理", "复用多模型列表对比", "待实现"),
    # 部署
    ("C38", "部署", "ONNX 导出", "torch.onnx.export", "待实现"),
    ("C39", "部署", "TorchScript 导出", "torch.jit.script/trace", "待实现"),
    # 可视化
    ("C40", "可视化", "批量推理结果 CSV 导出", "文件路径 + Top-1~Top-5 类别 + 置信度", "待实现"),
    ("C41", "可视化", "推理结果颜色标记", "正确分类绿色、错误分类红色", "待实现"),
    ("C42", "可视化", "类别置信度柱状图", "推理面板显示单图 Top-5 置信度 bar chart", "待实现"),
]

HEADER = ["序号", "模块", "功能", "详细说明", "状态"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "图像分类 Requirements"

# -- Styles --
header_font = Font(name="微软雅黑", bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
cell_font = Font(name="微软雅黑", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap_align = Alignment(wrap_text=True, vertical="center")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write header
for ci, h in enumerate(HEADER, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = header_font_white
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border

# Write data
for ri, row in enumerate(rows, 2):
    for ci, val in enumerate(row, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = cell_font
        c.border = thin_border
        if ci in (1, 5):  # 序号, 状态
            c.alignment = center_align
        else:
            c.alignment = wrap_align

# Column widths
col_widths = [8, 12, 22, 65, 10]
for ci, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:E{len(rows)+1}"

# --- Sheet 2: 对比汇总 ---
ws2 = wb.create_sheet("三任务对比")
ws2.append(["维度", "语义分割", "目标检测", "图像分类 (新)"])
ws2.append(["输出粒度", "像素级", "框级", "图级"])
ws2.append(["每图标注量", "数百万像素", "0~N 个框", "1 个标签"])
ws2.append(["标注成本", "极高", "高", "低"])
ws2.append(["训练框架", "torchvision", "Ultralytics YOLO", "torchvision + timm"])
ws2.append(["典型模型", "DeepLabV3", "YOLOv8", "ResNet / EfficientNet"])
ws2.append(["主要指标", "mIoU", "mAP@0.5", "Top-1 Accuracy"])
ws2.append(["损失函数", "CrossEntropy/Dice", "CIoU+DFL", "CrossEntropy / Focal"])
ws2.append(["推理输出", "mask 图", "框+分数", "类别+置信度"])

for row in ws2.iter_rows(min_row=1, max_row=9, min_col=1, max_col=4):
    for c in row:
        c.font = Font(name="微软雅黑", size=10)
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if c.row == 1:
            c.font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
            c.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 28
ws2.column_dimensions["C"].width = 28
ws2.column_dimensions["D"].width = 28
ws2.freeze_panes = "A2"

# --- Sheet 3: 实现阶段 ---
ws3 = wb.create_sheet("实现计划")
ws3.append(["阶段", "内容", "预估工时"])
ws3.append(["阶段 1: 核心代码", "classification/dataset.py + trainer.py + eval.py + predictor.classify()", "6-8h"])
ws3.append(["阶段 2: UI 集成", "main_window.py 分类面板 + 模式切换 + 推理结果展示", "8-10h"])
ws3.append(["阶段 3: 增强功能", "混淆矩阵 / ONNX导出 / 多标签 / 单元测试", "4-6h"])
ws3.append(["合计", "", "18-24h"])
for row in ws3.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3):
    for c in row:
        c.font = Font(name="微软雅黑", size=10)
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if c.row == 1:
            c.font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
            c.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")

ws3.column_dimensions["A"].width = 24
ws3.column_dimensions["B"].width = 65
ws3.column_dimensions["C"].width = 14
ws3.freeze_panes = "A2"

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Sheet 1: {len(rows)} requirements (C01-C{len(rows)})")
print(f"Sheet 2: 三任务对比 (9 rows)")
print(f"Sheet 3: 实现计划 (4 stages)")