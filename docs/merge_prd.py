"""Merge PRD_图像分类模块.xlsx into AI_Training_Platform_Requirements.xlsx, keep original filename."""
import os
import copy
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ORIGINAL = r"D:\01_D盘\01_Arthur\07_表格\24-需求表\AI_Training_Platform_Requirements.xlsx"
CLASSIFICATION = r"D:\01_D盘\01_Arthur\07_表格\24-需求表\PRD_图像分类模块.xlsx"
BACKUP = ORIGINAL.replace(".xlsx", "_backup.xlsx")

# Backup original
import shutil
shutil.copy2(ORIGINAL, BACKUP)
print(f"Backup saved: {BACKUP}")

# Load both workbooks
wb_src = openpyxl.load_workbook(CLASSIFICATION)
ws_src = wb_src["图像分类 Requirements"]

wb_dst = openpyxl.load_workbook(ORIGINAL)
ws_dst = wb_dst["Requirements"]

# Read source data (skip header row)
src_data = []
for r in range(2, ws_src.max_row + 1):
    row_vals = [ws_src.cell(r, c).value for c in range(1, ws_src.max_column + 1)]
    src_data.append(row_vals)

print(f"Source rows to append: {len(src_data)}")

# Read styles from the last data row of destination
last_row = ws_dst.max_row
styles = {}
for c in range(1, 6):
    cell = ws_dst.cell(last_row, c)
    styles[c] = {
        "font": copy.copy(cell.font),
        "border": copy.copy(cell.border),
        "alignment": copy.copy(cell.alignment),
        "fill": copy.copy(cell.fill),
    }

# Append each source row
for i, row_vals in enumerate(src_data):
    new_row = last_row + 1 + i
    for ci, val in enumerate(row_vals, 1):
        c = ws_dst.cell(row=new_row, column=ci, value=val)
        # Apply same style as last row
        c.font = copy.copy(styles[ci]["font"])
        c.border = copy.copy(styles[ci]["border"])
        c.alignment = copy.copy(styles[ci]["alignment"])
        # Only copy fill if it exists and not default
        if styles[ci]["fill"] and styles[ci]["fill"].start_color and styles[ci]["fill"].start_color.rgb:
            try:
                if styles[ci]["fill"].start_color.rgb != "00000000":
                    c.fill = copy.copy(styles[ci]["fill"])
            except:
                pass

print(f"Destination now has {ws_dst.max_row} rows (1 header + {ws_dst.max_row - 1} data)")

# Update auto-filter to cover all rows
ws_dst.auto_filter.ref = f"A1:E{ws_dst.max_row}"

# Save merged
wb_dst.save(ORIGINAL)
print(f"Merged saved: {ORIGINAL}")
print(f"Original 105 requirements + 42 classification = {ws_dst.max_row - 1} total")